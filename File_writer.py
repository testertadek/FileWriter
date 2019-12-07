import openpyxl

path = "C:\\tadek\\2019\\python\\Projekty\\FileWriter\\excel_file\\tadek.xlsx" #input file

#read object
work_book_read = openpyxl.load_workbook(path)
sheet_obj_read = work_book_read.active
number_of_rows_read = sheet_obj_read.max_row
number_of_columns_read = sheet_obj_read.max_column #not used variable

#write object
work_book_write = openpyxl.Workbook()
sheet_obj_write = work_book_write.active

for row in range(number_of_rows_read):
        row_number  = row +1
        #data in one row in read object
        all_data_in_row = sheet_obj_read.cell(row = row_number, column=1)

        #list of items from one row in read object
        list_of_items_in_row = str(all_data_in_row.value).split(",")

        for column in range(len(list_of_items_in_row)):
            column_number = column+1
            #identify column in write object and wtite single value from list of items  from read ongejct
            value_to_write = sheet_obj_write.cell(row = row_number, column=column_number)
            value_to_write.value = list_of_items_in_row[column]

#save write oject
work_book_write.save("C:\\tadek\\2019\\python\\Projekty\\FileWriter\\excel_file\\NOwyPlik.xlsx")
